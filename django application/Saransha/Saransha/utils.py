import os
import time
import random
import re
from datetime import datetime
from typing import List, Tuple, Optional

import pandas as pd
import numpy as np
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

MAX_PROFILES = 10
MAX_PUBLICATIONS = 100
REQUEST_DELAY = 2

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Cache-Control': 'max-age=0',
}


def get_publications_from_profile(profile_url: str, timeout: int = 30, max_publications: int = 100) -> List[dict]:
    if "user=" not in profile_url:
        raise ValueError("Invalid Google Scholar profile URL")

    user_id = profile_url.split("user=")[1].split("&")[0]
    
    print(f"Fetching author data for user ID: {user_id}")
    print("  - Sending request to Google Scholar...")
    
    base_url = f"https://scholar.google.com/citations?user={user_id}&hl=en&cstart=0&pagesize=100"
    
    try:
        time.sleep(random.uniform(2, 4))
        
        session = requests.Session()
        session.get("https://scholar.google.com", headers=HEADERS, timeout=10)
        time.sleep(1)
        
        response = session.get(base_url, headers=HEADERS, timeout=timeout)
        
        if response.status_code == 429:
            raise Exception("HTTP Error 429 - Too Many Requests. Google Scholar is blocking your IP.")
        
        if response.status_code != 200:
            raise Exception(f"HTTP Error {response.status_code}")
        
        html = response.text
        print("  - Response received, parsing data...")
        
        if "CAPTCHA" in html or "captcha" in html:
            raise Exception("Google Scholar is showing CAPTCHA. Please try using a VPN.")
        
        soup = BeautifulSoup(html, 'html.parser')
        
        main_author = 'Unknown Author'
        name_elem = soup.find('div', id='gsc_prf_in')
        if name_elem:
            main_author = name_elem.text.strip()
        
        if main_author == 'Unknown Author':
            title_tag = soup.find('title')
            if title_tag and ' - Google Scholar' in title_tag.text:
                main_author = title_tag.text.replace(' - Google Scholar', '').strip()
        
        print(f"  - Author found: {main_author}")
        
        pub_rows = soup.find_all('tr', class_='gsc_a_tr')
        
        if not pub_rows:
            debug_file = f"debug_scholar_{user_id}.html"
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(html)
            print(f"  - Debug HTML saved to {debug_file}")
            raise Exception("No publications found. Check debug HTML file.")
        
        print(f"  - Found {len(pub_rows)} publication rows")
        
        today = datetime.now().strftime("%Y-%m-%d")
        publications = []
        
        for idx, row in enumerate(pub_rows[:max_publications]):
            try:
                title = ''
                title_elem = row.find('a', class_='gsc_a_at')
                if title_elem:
                    title = title_elem.text.strip()
                
                if not title or len(title) < 5:
                    continue
                
                gray_texts = row.find_all('div', class_='gs_gray')
                co_authors = main_author
                venue = 'N/A'
                
                if len(gray_texts) >= 1:
                    co_authors = gray_texts[0].text.strip()
                if len(gray_texts) >= 2:
                    venue = gray_texts[1].text.strip()
                
                journal = 'N/A'
                conference = 'N/A'
                venue_lower = venue.lower()
                
                if any(word in venue_lower for word in ['conference', 'proceedings', 'symposium', 'workshop']):
                    conference = venue
                else:
                    journal = venue if venue != 'N/A' else 'N/A'
                
                year = None
                year_elem = row.find('span', class_='gsc_a_h')
                if year_elem:
                    year_text = year_elem.text.strip()
                    if year_text.isdigit():
                        year = int(year_text)
                
                citations = 0
                cite_elem = row.find('a', class_='gsc_a_ac')
                if cite_elem:
                    cite_text = cite_elem.text.strip()
                    if cite_text.isdigit():
                        citations = int(cite_text)
                
                publications.append({
                    'Main Author': main_author,
                    'Title': title,
                    'Journal': journal,
                    'conference': conference,
                    'Year': year,
                    'Publication Type': 'conference' if conference != 'N/A' else 'article',
                    'Cited by': citations,
                    'co_author': co_authors,
                    'Last Search Date': today
                })
                
            except Exception as e:
                continue
        
        if not publications:
            raise Exception("Could not parse any publications from the page.")
        
        print(f"Successfully processed {len(publications)} publications for {main_author}")
        return publications
        
    except requests.exceptions.Timeout:
        raise Exception("Request timed out.")
    except requests.exceptions.RequestException as e:
        raise Exception(f"Network error: {str(e)}")


def get_publications_safe(profile_url: str, timeout: int = 30, max_publications: int = 100) -> List[dict]:
    max_retries = 3
    
    for attempt in range(max_retries):
        try:
            return get_publications_from_profile(profile_url, timeout, max_publications)
        except Exception as e:
            print(f"  - Attempt {attempt + 1} failed: {e}")
            if "429" in str(e) or "CAPTCHA" in str(e):
                return []
            if attempt < max_retries - 1:
                wait_time = (attempt + 1) * 3
                print(f"  - Waiting {wait_time} seconds before retry...")
                time.sleep(wait_time)
    
    return []


def process_profiles_from_excel(file_path: str, output_file: str):
    if not os.path.exists(file_path):
        raise FileNotFoundError("Excel file not found")

    df = pd.read_excel(file_path)

    if 'Profile URL' not in df.columns:
        raise ValueError("Excel must contain 'Profile URL' column")

    all_publications = []
    profile_urls = df['Profile URL'].dropna().tolist()[:MAX_PROFILES]
    
    for idx, profile_url in enumerate(profile_urls):
        try:
            print(f"\n[INFO] Processing profile {idx + 1}/{len(profile_urls)}")
            print(f"URL: {profile_url}")
            
            pubs = get_publications_safe(profile_url)
            
            if pubs:
                all_publications.extend(pubs)
                print(f"[SUCCESS] Got {len(pubs)} publications")
            
            if idx < len(profile_urls) - 1:
                delay = random.uniform(5, 8)
                print(f"  - Waiting {delay:.1f}s before next profile...")
                time.sleep(delay)
                
        except Exception as e:
            print(f"[ERROR] {profile_url} -> {e}")

    if not all_publications:
        raise ValueError("No publications collected.")

    result_df = pd.DataFrame(all_publications)
    result_df.to_excel(output_file, index=False)
    print(f"\n[SUCCESS] Saved {len(all_publications)} publications to: {output_file}")
    
    return result_df


def generate_author_summary(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df['Cited by'] = pd.to_numeric(df['Cited by'], errors='coerce').fillna(0)

    if 'Journal' in df.columns:
        summary = df.groupby('Main Author').agg(
            publication=('Title', 'count'),
            journal=('Journal', lambda x: (x != 'N/A').sum()),
            total_citations=('Cited by', 'sum')
        ).reset_index()
    else:
        summary = df.groupby('Main Author').agg(
            publication=('Title', 'count'),
            total_citations=('Cited by', 'sum')
        ).reset_index()
        summary['journal'] = 0

    return summary.sort_values(by='total_citations', ascending=False)


def generate_publication_summary(df: pd.DataFrame):
    df = df.dropna(subset=['Year'])
    if df.empty:
        return [], {}
        
    df['Year'] = df['Year'].astype(int)
    years = list(range(df['Year'].min(), df['Year'].max() + 1))
    authors = df['Main Author'].unique()

    result = {a: [0] * len(years) for a in authors}

    for _, row in df.iterrows():
        year_idx = years.index(row['Year'])
        result[row['Main Author']][year_idx] += 1

    return list(map(str, years)), result


def update_publication_details(file_path, author, title, journal, conference, year):
    wb = load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == author and row[1].value == title:
            row[2].value = journal
            row[3].value = conference
            row[4].value = year
            wb.save(file_path)
            return "Update successful"

    return "Publication not found"


def build_publication_context(df: pd.DataFrame, query: str, top_k: int = 4) -> Tuple[str, List[dict]]:
    if df.empty or not query:
        return "", []

    query = query.lower()
    scored = []

    for _, row in df.iterrows():
        score = 0
        for field in ['Title', 'Main Author', 'Journal', 'conference']:
            if query in str(row.get(field, '')).lower():
                score += 3
        if score > 0:
            scored.append((score, row))

    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[:top_k]

    context, refs = [], []

    for _, row in top:
        context.append(
            f"Title: {row['Title']}\n"
            f"Author: {row['Main Author']}\n"
            f"Journal: {row['Journal']}\n"
            f"Year: {row['Year']}\n"
            f"Cited by: {row['Cited by']}\n"
        )
        refs.append({
            'title': row['Title'],
            'author': row['Main Author'],
            'year': row['Year']
        })

    return "\n---\n".join(context), refs


def load_and_filter_excel(
    file_path,
    sheet_name='Sheet1',
    columns=None,
    column_name=None,
    valid_names=None,
    cited_by_sort_order=None,
    year_range=None
):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} does not exist")

    df = pd.read_excel(file_path, sheet_name=sheet_name)

    if columns:
        columns = [c for c in columns if c in df.columns]
        df = df[columns]

    if column_name and valid_names and column_name in df.columns:
        df = df[df[column_name].isin(valid_names)]

    if cited_by_sort_order:
        if cited_by_sort_order in ['asc', 'desc'] and 'Cited by' in df.columns:
            df['Cited by'] = pd.to_numeric(df['Cited by'], errors='coerce').fillna(0)
            df = df.sort_values(by='Cited by', ascending=(cited_by_sort_order == 'asc'))
        elif cited_by_sort_order == 'Date' and 'Year' in df.columns:
            df = df.sort_values(by='Year')

    if year_range and len(year_range) == 2 and 'Year' in df.columns:
        df = df[(df['Year'] >= year_range[0]) & (df['Year'] <= year_range[1])]

    return df


SCOPUS_URL = "https://api.elsevier.com/content/serial/title"

def check_scopus_index_for_df(df: pd.DataFrame, api_key: str) -> pd.DataFrame:
    if 'Journal' not in df.columns:
        raise ValueError("DataFrame must contain 'Journal' column")

    headers = {'Accept': 'application/json', 'X-ELS-APIKey': api_key}

    def is_indexed(journal):
        if not journal or journal == 'N/A':
            return False
        try:
            r = requests.get(SCOPUS_URL, headers=headers, params={'title': journal}, timeout=10)
            return r.status_code == 200 and bool(r.json().get('serial-metadata-response'))
        except:
            return None

    df['Scopus Indexed'] = df['Journal'].apply(is_indexed)
    return df